
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from pathlib import Path
from typing import Dict, Any, List
import logging
import csv
import io

logger = logging.getLogger(__name__)

def read_table(path: str, sheet_name: str = None) -> List[Dict[str, Any]]:
    """Read first sheet or specified sheet and return list of rows as dicts (header -> value)."""
    wb = load_workbook(filename=path, data_only=True)
    sh = wb[sheet_name] if sheet_name else wb.active
    rows = list(sh.rows)
    if not rows:
        return []
    headers = [cell.value for cell in rows[0]]
    data = []
    for row in rows[1:]:
        item = {}
        for h, c in zip(headers, row):
            item[h] = c.value
        data.append(item)
    return data

def write_summary(path: str, summary: Dict[str, Any], out_path: str = None) -> str:
    """Write summary dict into a new sheet named 'Summary' in the same workbook or to a new workbook.

    Returns the path written.
    """
    p = Path(path)
    wb = load_workbook(filename=path)
    # remove existing Summary sheet if present
    if 'Summary' in wb.sheetnames:
        std = wb['Summary']
        wb.remove(std)
    sh = wb.create_sheet('Summary')
    sh.append(['Metric', 'Value'])
    for k, v in summary.items():
        sh.append([k, str(v)])
    target = out_path or str(p)
    wb.save(target)
    logger.info('Wrote Summary to %s', target)
    return target

def compute_summary(rows: List[Dict[str, Any]], numeric_fields: List[str] = None) -> Dict[str, Any]:
    """Compute simple summary: row_count, for numeric fields compute sum and average."""
    numeric_fields = numeric_fields or []
    summary = {}
    summary['row_count'] = len(rows)
    for field in numeric_fields:
        vals = [r.get(field) for r in rows if isinstance(r.get(field), (int, float))]
        total = sum(vals) if vals else 0
        avg = total / len(vals) if vals else 0
        summary[f'{field}_total'] = total
        summary[f'{field}_average'] = round(avg, 2)
    return summary

def rows_to_csv_bytes(rows, headers):
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    for r in rows:
        writer.writerow([r.get(h, '') for h in headers])
    return buf.getvalue().encode('utf-8')
